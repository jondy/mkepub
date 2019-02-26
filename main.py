#! /usr/bin/env python
# -*- coding: utf-8 -*-
#

import os
import sys

from glob import glob

from PyQt5.QtCore import Qt, QDir, QSettings, QUrl, QFileInfo, \
    QThread, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QFont, QDesktopServices
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, \
    QFileDialog, QTableWidgetItem

from ui_main import Ui_MainWindow

COL_FILENAME = 0
COL_STATUS = 1

# from transform import process_file, save_result

from openpyxl import load_workbook


def save_result(filelist, result):
    filename = 'upload-yyyy-mm-dd-hh-mi-ss.xlsx'
    template = os.path.join(os.path.dirname(__file__), 'upload.xltx')
    wb = load_workbook(template)
    wb.template = False

    start_row = 4
    filename_col = 'A'
    page_col = 'C'
    total_col = 'D'
    non_blank_col = 'E'
    han_col = 'F'
    price_col = 'G'
    price = 450.0
    money_col = 'H'
    money_formula = '=F{row}/1000*G{row}'

    ws = wb.active
    row = start_row
    for name, counter in result:
        if counter is None:
            continue
        page, total, han, blank = counter
        rs = str(row)
        ws[filename_col + rs] = os.path.basename(name)
        ws[page_col + rs] = page
        ws[total_col + rs] = total
        ws[non_blank_col + rs] = total - blank
        ws[han_col + rs] = han
        ws[price_col + rs] = price
        ws[money_col + rs] = money_formula.format(row=row)
        row += 1

    sum_formula = '=sum({col}%d:{col}%d)' % (start_row, row-1)
    rs = str(row)
    ws[filename_col + rs] = '总计'
    ws[page_col + rs] = sum_formula.format(col=page_col)
    ws[total_col + rs] = sum_formula.format(col=total_col)
    ws[non_blank_col + rs] = sum_formula.format(col=non_blank_col)
    ws[han_col + rs] = sum_formula.format(col=han_col)
    ws[money_col + rs] = sum_formula.format(col=money_col)

    wb.save(filename)
    wb.close()

    return filename


def process_file(filename):
    pass


class EpubWorker(QThread):

    fileStart = pyqtSignal(int)
    fileEnd = pyqtSignal(int, dict)

    def __init__(self, filelist, parent=None):
        super(EpubWorker, self).__init__(parent)
        self.filelist = filelist
        self.request_stop = 0

    def run(self):
        row = 0
        for filename in self.filelist:
            if self.request_stop:
                break
            self.fileStart.emit(row)
            result = process_file(filename)
            self.fileEnd.emit(row, result)
            row += 1


class MainWindow(QMainWindow, Ui_MainWindow):

    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)

        self._settings = QSettings('Dashingsoft', 'Word-Counter')
        self._lastPath = self._settings.value('lastPath', QDir.currentPath())

        self.actionSelectDirectory.triggered.connect(self.selectDirectory)
        self.actionSelectFiles.triggered.connect(self.selectFiles)
        self.actionStart.triggered.connect(self.startTransform)
        self.actionStop.triggered.connect(self.stopTransform)
        self.actionUpload.triggered.connect(self.uploadFiles)
        self.actionAbout.triggered.connect(self.about)

        self._filelist = []
        self._result = []
        self._worker = None

    def _setLastPath(self, path):
        self._lastPath = path
        self._settings.setValue('lastPath', path)

    def _initFileList(self, filelist):
        w = self.tableWidget
        for i in range(w.rowCount()):
            w.removeRow(0)

        row = 0
        for filename in filelist:
            w.insertRow(row)
            w.setItem(row, 0, QTableWidgetItem(os.path.basename(filename)))
            row += 1
        w.horizontalHeader().setSectionResizeMode(3)

    def selectDirectory(self):
        directory = QFileDialog.getExistingDirectory(self, '选择目录',
                                                     self._lastPath)
        if directory:
            self._setLastPath(directory)
            filelist = glob(os.path.join(directory, '*.txt'))
            if filelist:
                self._initFileList(filelist)
            else:
                QMessageBox.information(self, '选择目录', '在选择的目录下面没有任何 .txt 文件')

    def selectFiles(self):
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(
            self, '选择文件',
            self._lastPath, '文本文件 (*.txt)',
            options=options)

        if files:
            self._setLastPath(os.path.dirname(files[0]))
            self._initFileList(files)

    def about(self):
        QMessageBox.about(self, '关于', '红云电子书制作工具 v0.1a1')

    def startTransform(self):
        if self._filelist:
            self._transformFiles(self._filelist)

    def stopTransform(self):
        if self._worker:
            self._worker.request_stop = 1

    def uploadFiles(self):
        pass

    @pyqtSlot(int)
    def handleFileStart(self, row):
        w = self.tableWidget
        w.setCurrentCell(row, 0)
        w.setItem(row, 4, QTableWidgetItem('正在统计...'))

    @pyqtSlot(int, dict)
    def handleFileEnd(self, row, result):
        w = self.tableWidget
        if result is None:
            w.item(row, COL_STATUS).setText('转换失败')
            w.item(row, COL_STATUS).setBackground(Qt.red)
        else:
            w.item(row, COL_STATUS).setText('转换完成')
            w.item(row, COL_STATUS).setBackground(Qt.lightGray)
        self._result.append(result)

    @pyqtSlot()
    def handleWorkerFinished(self):
        if not self._worker.request_stop:
            self.tableWidget.setCurrentItem(None)
            output = save_result(self._filelist, self._result)
            msg = '文件转换完成，上传列表已经保存在文件: %s, 是否查看?'
            reply = QMessageBox.question(self, '转换结束', msg % output)
            if reply == QMessageBox.Yes:
                QDesktopServices.openUrl(
                    QUrl.fromLocalFile(QFileInfo(output).absoluteFilePath()))
        self._worker = None

    def _transformFiles(self, filelist):
        self._filelist = filelist
        worker = EpubWorker(filelist, parent=self)
        worker.fileStart.connect(self.handleFileStart)
        worker.fileEnd.connect(self.handleFileEnd)
        worker.finished.connect(self.handleWorkerFinished)
        worker.start()
        self._worker = worker


def main():
    app = QApplication(sys.argv)
    font = QFont('宋体')
    pointsize = font.pointSize()
    font.setPixelSize(pointsize*90/72)
    app.setFont(font)

    window = MainWindow()
    window.show()

    app.exec_()


if __name__ == '__main__':
    main()
