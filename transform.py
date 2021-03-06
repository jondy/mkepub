#! /usr/bin/env python3
# -*- coding: utf-8 -*-
#

import argparse
import logging
import os
import subprocess
import sys
import uuid

from ebooklib import epub
from openpyxl import load_workbook

from readers import find_reader

version_info = '0.1a0'

UPLOAD_PATH = '/www/wwwroot/www.yancloud.red/Uploads/bookTemp'


def upload_file(filename):
    cmdlist = ['ftp']
    cmdlist.append(filename)
    p = subprocess(cmdlist,
                   stdout=subprocess.PIPE,
                   stderr=subprocess.STDOUT)
    output, _ = p.communicate()


def save_result(filelist, result, filename=None):
    if filename is None:
        filename = 'upload-epub.xlsx'
    template = os.path.join(os.path.dirname(__file__), 'upload.xltx')
    wb = load_workbook(template)
    wb.template = False

    start_row = 3
    filename_col = 'A'
    author_col = 'D'
    author_info_col = 'E'
    isbn_col = 'F'
    date_col = 'G'
    publisher_col = 'H'
    price_col = 'I'
    intro_col = 'J'
    path_col = 'K'

    def format_author(a):
        return ','.join(a.split())

    def format_date(a):
        for t in ('年', '月', '日', '/'):
            a = a.replace(t, '-')
        s = [x.strip() for x in a.split('-') if x.strip()]
        if len(s) == 2:
            s.append('1')
        return '-'.join(s)

    ws = wb.active
    row = start_row
    for name, meta in zip(filelist, result):
        rs = str(row)
        name = os.path.basename(name).rsplit('.', 1)[0]
        ws[filename_col + rs] = meta.get('title', name)
        ws[author_col + rs] = format_author(meta.get('author', ''))
        ws[author_info_col + rs] = meta.get('author_info')
        ws[isbn_col + rs] = meta.get('ISBN')
        ws[date_col + rs] = format_date(meta.get('date', ''))
        ws[publisher_col + rs] = meta.get('publisher')
        ws[price_col + rs] = meta.get('price')
        ws[intro_col + rs] = meta.get('intro')
        ws[path_col + rs] = name + '.epub'
        row += 1

    wb.save(filename)
    wb.close()

    return filename


def process_file(filename, output='output'):
    logging.info('Processing %s...', filename)
    reader = find_reader(filename)
    if reader is None:
        raise Exception('不支持的文件类型')

    reader.open(filename)

    book = epub.EpubBook()
    book.FOLDER_NAME = 'OEBPS'

    style = '''body { qrfullpage:1; text-align:center; }
               img { max-width: 80% }'''
    cover_css = epub.EpubItem(uid="style_cover",
                              file_name="cover.css",
                              media_type="text/css",
                              content=style)
    book.add_item(cover_css)

    path = os.path.dirname(__file__)
    with open(os.path.join(path, 'templates', 'default.css')) as f:
        default_css = epub.EpubItem(uid="style_default",
                                    file_name="../Styles/default.css",
                                    media_type="text/css",
                                    content=f.read())
        book.add_item(default_css)

    meta = reader.get_metadata()
    book.set_identifier(meta.get('ISBN', str(uuid.uuid4())))

    name = os.path.splitext(os.path.basename(filename))[0]
    book.set_title(meta.get('title', name))
    book.set_language('zh')

    author = meta.get('author')
    if author:
        book.add_author(author)

    cover = reader.get_cover()
    if cover:
        book.set_cover('Images/coverpage.jpg', open(cover, 'rb').read())
        book.get_item_with_id('cover').add_item(cover_css)
        book.toc = [epub.Link('cover.xhtml', '封面', 'cover')]
    else:
        book.toc = []

    css_items = []
    for item in reader.stylesheets():
        book.add_item(item)
        css_items.append(item)

    for item in reader.contents():
        if isinstance(item, epub.EpubHtml):
            item.add_item(default_css)
            for css in css_items:
                item.add_item(css)
        book.add_item(item)

    for item in reader.images():
        book.add_item(item)

    # sec = None
    # for item in reader.get_toc():
    #     n, p = item
    #     if isinstance(p, str):
    #         if sec is not None:
    #             book.toc.append(sec)
    #         s = epub.Section(p)
    #         sec = s, []
    #     else:
    #         if sec[0].href == '':
    #             sec[0].href = p.get_name()
    #         sec[1].append(p)
    # if sec is not None:
    #     book.toc.append(sec)

    toc = reader.get_toc()
    if toc is None:
        raise Exception('不正确的章节结构')
    book.toc.extend(toc)

    reader.close()

    book.add_item(epub.EpubNcx())
    book.add_item(epub.EpubNav())

    book.spine = ['cover', 'nav']
    book.spine.extend(list(book.get_items_of_type(9))[1:-1])

    if not os.path.exists(output):
        os.makedirs(output)
    epub.write_epub(os.path.join(output, name + '.epub'), book)
    return meta


def main(args):
    parser = argparse.ArgumentParser(
        prog='mkepub',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description='Make epub from pdf or text file',
    )
    parser.add_argument('-v', '--version', action='version',
                        version=version_info)
    parser.add_argument('-q', '--silent', action='store_true',
                        help='Suppress all normal output')

    parser.add_argument('-O', '--output', default='output', metavar='PATH')
    parser.add_argument('-c', '--cover', metavar='IMAGE',
                        help='Filename of cover image]')
    parser.add_argument('-t', '--template', metavar='PATH',
                        help='Template path')
    parser.add_argument('filenames', nargs='+', help='Source filenames')

    args = parser.parse_args(args)
    if args.silent:
        logging.getLogger().setLevel(100)

    for filename in args.filenames:
        process_file(filename, args.output)


def main_entry():
    logging.basicConfig(
        level=logging.INFO,
        format='%(levelname)-8s %(message)s',
    )
    try:
        main(sys.argv[1:])
    except Exception as e:
        if sys.flags.debug:
            raise
        logging.error('%s', e)
        sys.exit(1)


if __name__ == '__main__':
    main_entry()
    # main(['test/examples/解读延安精神.txt'])
